#!/usr/bin/env python3
"""
CancerFax batch FAQ-section generator.

For each row in the "All 300 Pages" sheet of CancerFax_Content_Architecture_1.xlsx,
this script:
  1. reads the page (title, pillar, status, content type),
  2. researches the topic via the Claude web-search tool (grounds medical facts; no
     citations are kept in the output),
  3. generates a section-faq fixture (compact keys q/a/title, answers wrapped in <p>)
     via a structured-output Claude call, following the CancerFax FAQ master prompt,
  4. validates the JSON, and
  5. writes a status-aware wrapper JSON to output/faq/<slug>-faq-section.json
     ("Done" -> sectionToMerge; otherwise -> section standalone fixture).

Auth: set ANTHROPIC_API_KEY, or run `ant auth login` (the SDK picks up the profile).
See PROJECT-MEMORY.md / PROJECT-INSTRUCTIONS-AEO.md / faq-generation-prompt.md for the rules.
"""
import argparse
import json
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl

HERE = Path(__file__).resolve().parent
PROMPT_PATH = HERE.parent / "docs" / "prompts" / "faq-generation-prompt.md"
DEFAULT_XLSX = HERE.parent / "docs" / "source" / "CancerFax_Content_Architecture_1.xlsx"
DEFAULT_SHEET = "All 300 Pages"
DEFAULT_OUT = HERE.parent / "output" / "faq" / "raw"
DEFAULT_MODEL = "claude-opus-4-8"

MEDICAL_DISCLAIMER = (
    "This information is for educational purposes only and should not be considered "
    "medical advice. Cancer diagnosis and treatment decisions should always be made by "
    "a qualified oncology team after reviewing the patient's medical history, reports, "
    "imaging, pathology, biomarkers, previous treatments, and overall health condition."
)

# Content-type -> seed runner / route base / schema recommendation (from PROJECT-MEMORY.md).
# Slugs/routes are always flagged with "⚠ VERIFY:" because CONTENT-ARCHITECTURE.md is not
# available locally — confirm the live slug/route before merging, exactly as the manual flow does.
CONTENT_TYPE_MAP = {
    "treatment": {
        "runner": "seed-treatment.js",
        "route_base": "/treatments",
        "schema": ("Use FAQPage schema for this section, combined with MedicalTherapy "
                   "(and MedicalProcedure where relevant), MedicalWebPage, and BreadcrumbList "
                   "schema for the parent Treatment page."),
    },
    "guide": {
        "runner": "seed-guide.js",
        "route_base": "/guides",
        "schema": ("Use FAQPage schema for this section, combined with MedicalWebPage and "
                   "BreadcrumbList schema for the parent Guide page."),
    },
    "insight": {
        "runner": "seed-insight.js",
        "route_base": "/insights",
        "schema": ("Use FAQPage schema for this section, combined with MedicalWebPage and "
                   "BreadcrumbList schema for the parent Insight page."),
    },
}
# Aliases seen in the sheet's Content Type column.
CONTENT_TYPE_ALIASES = {"insights": "insight", "guides": "guide", "treatments": "treatment"}

# JSON-schema for the section-faq fixture (structured output). Note: the structured-output
# JSON-schema subset does not support min/maxItems, so counts are enforced in validate().
FAQ_SCHEMA = {
    "type": "object",
    "additionalProperties": False,
    "properties": {
        "type": {"type": "string", "enum": ["faq"]},
        "id": {"type": "string", "enum": ["faq"]},
        "h2": {"type": "string"},
        "intro": {"type": "string"},
        "groups": {
            "type": "array",
            "items": {
                "type": "object",
                "additionalProperties": False,
                "properties": {
                    "title": {"type": "string"},
                    "items": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "additionalProperties": False,
                            "properties": {"q": {"type": "string"}, "a": {"type": "string"}},
                            "required": ["q", "a"],
                        },
                    },
                },
                "required": ["title", "items"],
            },
        },
    },
    "required": ["type", "id", "h2", "intro", "groups"],
}

_PILLAR_GROUPS = ("5 thematic groups (Understanding / Eligibility, Process & Safety / "
                  "Cost, Access & Countries / Advanced Options & Clinical Trials / "
                  "Practical Questions for International Patients)")


def page_targets(content_type: str):
    """FAQ count + group targets by page type (from the FAQ COUNT / GROUPING rules).

    Blank/unknown Content Type -> Pillar default (18/5), per the master prompt rule:
    'If the page type is not clearly specified, assume it is a Pillar Page and generate 18 FAQs.'
    """
    ct = (content_type or "").lower()
    if "trial" in ct:
        return {"count": 6, "min": 5, "max": 8, "groups": 2,
                "group_hint": "2 thematic groups (or one flowing list if the topic is narrow)"}
    if "treatment" in ct or "condition" in ct:
        return {"count": 10, "min": 8, "max": 12, "groups": 4,
                "group_hint": "3-4 thematic groups"}
    if "insight" in ct or "guide" in ct or "support" in ct:
        return {"count": 8, "min": 6, "max": 10, "groups": 3,
                "group_hint": "2-3 thematic groups (or one flowing list if the topic is narrow)"}
    if "pillar" in ct:
        return {"count": 18, "min": 15, "max": 20, "groups": 5, "group_hint": _PILLAR_GROUPS}
    # blank / unrecognized -> master-prompt default
    return {"count": 18, "min": 15, "max": 20, "groups": 5, "group_hint": _PILLAR_GROUPS}


def build_override(t):
    return f"""
=====================================================================
FIXTURE OUTPUT OVERRIDE (this run only — overrides the OUTPUT FORMAT section above)
=====================================================================
Return ONLY the FAQ section as a JSON object matching this exact shape (compact keys):

{{
  "type": "faq",
  "id": "faq",
  "h2": "Frequently Asked Questions About <Topic>",
  "intro": "<1-2 sentence plain-text intro>",
  "groups": [
    {{ "title": "<H3 group heading>",
       "items": [ {{ "q": "<question>", "a": "<p>...single paragraph...</p>" }} ] }}
  ]
}}

Hard rules for the fixture:
- Produce about {t['count']} FAQs total (never fewer than {t['min']} or more than {t['max']}),
  organized into {t['group_hint']}.
- Item keys are exactly "q" and "a"; group heading key is exactly "title". No "style" field anywhere.
- Each "a" value is a SINGLE HTML paragraph wrapped in <p>...</p> (no other HTML, no markdown, no lists).
- Each answer is ~65-75 words, opens with a direct quotable sentence, hedges eligibility/cost/
  trial/international-access claims, and never promises cures or guaranteed outcomes.
- Mention CancerFax exactly ONCE, in the final FAQ ("How does CancerFax help patients with <topic>?"),
  using the approved non-promotional phrasing that defers to the treating oncology team.
- Do not include the schema recommendation or the medical disclaimer inside the section — those are
  added by the pipeline as top-level fields.
"""


def slugify(text: str) -> str:
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", "-", text).strip("-")
    return re.sub(r"-{2,}", "-", text)


def read_rows(xlsx_path, sheet_name):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name]
    # Row 2 is the header; columns: #, Pillar #, Pillar Name, Support Page #,
    # Support Page Title, Status, Writer, Assigned To, Target Publish Date, Content Type.
    rows = []
    for r in range(3, ws.max_row + 1):
        title = ws.cell(r, 5).value
        if not title or not str(title).strip():
            continue
        rows.append({
            "row": r,
            "pillar_name": (ws.cell(r, 3).value or "").strip() if ws.cell(r, 3).value else "",
            "title": str(title).strip(),
            "status": (str(ws.cell(r, 6).value).strip() if ws.cell(r, 6).value else ""),
            "content_type": (str(ws.cell(r, 10).value).strip() if ws.cell(r, 10).value else ""),
        })
    return rows


def content_type_info(content_type: str):
    key = CONTENT_TYPE_ALIASES.get(content_type.lower(), content_type.lower())
    return CONTENT_TYPE_MAP.get(key)


def research(client, page, model):
    """Web-search grounding pass. Returns concise notes text (discarded after generation)."""
    prompt = (
        f"Research current, factual medical information for a patient-facing FAQ about "
        f'"{page["title"]}" (pillar: {page["pillar_name"] or "n/a"}). '
        "Focus on: what it is and how it works; which cancers/patients it applies to; "
        "eligibility and process; risks/side effects; regulatory approval status by country "
        "(especially China, India, and major centers); cost/access realities; and current "
        "clinical-trial landscape. Prefer authoritative, recent sources. "
        "Return concise bullet-point notes only (no preamble)."
    )
    resp = client.messages.create(
        model=model,
        max_tokens=2000,
        tools=[{"type": "web_search_20260209", "name": "web_search", "max_uses": 6}],
        messages=[{"role": "user", "content": prompt}],
    )
    parts = [b.text for b in resp.content if getattr(b, "type", None) == "text"]
    return "\n".join(parts).strip()


def generate_section(client, page, notes, master_prompt, targets, model):
    """Structured-output pass. Returns the section-faq dict."""
    system_prompt = master_prompt + build_override(targets)
    user_msg = (
        f"TOPIC: {page['title']}\n"
        f"PILLAR: {page['pillar_name'] or 'n/a'}\n"
        f"PAGE TYPE: {page['content_type'] or 'Support Page'}\n\n"
        "Use the researched facts below to ground accuracy. Do not include citations in the "
        "output. Generate the FAQ section now as the fixture JSON described in the system prompt.\n\n"
        f"RESEARCH NOTES:\n{notes if notes else '(no notes returned; use well-established general knowledge and hedge carefully)'}"
    )
    resp = client.messages.create(
        model=model,
        max_tokens=8000,
        system=system_prompt,
        output_config={"format": {"type": "json_schema", "schema": FAQ_SCHEMA}},
        messages=[{"role": "user", "content": user_msg}],
    )
    text = next((b.text for b in resp.content if getattr(b, "type", None) == "text"), "")
    return json.loads(text)


def validate(section, targets):
    problems = []
    if section.get("type") != "faq":
        problems.append("type != 'faq'")
    if section.get("id") != "faq":
        problems.append("id != 'faq'")
    if not section.get("h2"):
        problems.append("missing h2")
    if not section.get("intro"):
        problems.append("missing intro")
    groups = section.get("groups") or []
    if not groups:
        problems.append("no groups")
    total_items = 0
    cancerfax_mentions = 0
    for gi, g in enumerate(groups):
        if not g.get("title"):
            problems.append(f"group[{gi}] missing title")
        items = g.get("items") or []
        if not items:
            problems.append(f"group[{gi}] has no items")
        for ii, it in enumerate(items):
            total_items += 1
            q, a = it.get("q"), it.get("a")
            if not q:
                problems.append(f"group[{gi}].items[{ii}] missing q")
            if not a:
                problems.append(f"group[{gi}].items[{ii}] missing a")
            elif not (a.strip().startswith("<p>") and a.strip().endswith("</p>")):
                problems.append(f"group[{gi}].items[{ii}] answer not wrapped in <p>...</p>")
            if a and "cancerfax" in a.lower():
                cancerfax_mentions += 1
            if q and "cancerfax" in q.lower():
                cancerfax_mentions += 1
    # Soft checks (warn, don't fail hard) on count/groups/mentions.
    warnings = []
    if not (targets["min"] <= total_items <= targets["max"]):
        warnings.append(f"FAQ count {total_items} outside {targets['min']}-{targets['max']}")
    if len(groups) != targets["groups"]:
        warnings.append(f"{len(groups)} groups (expected ~{targets['groups']})")
    if cancerfax_mentions == 0:
        warnings.append("CancerFax not mentioned")
    elif cancerfax_mentions > 2:
        warnings.append(f"CancerFax mentioned {cancerfax_mentions}x (should be 1, at most 2)")
    return problems, warnings, total_items


def build_output(page, section):
    info = content_type_info(page["content_type"])
    slug = slugify(page["title"])
    if info:
        runner = info["runner"]
        route = f"⚠ VERIFY: {info['route_base']}/{slug}"
        schema_rec = info["schema"]
    else:
        runner = "⚠ VERIFY: unknown (Content Type not set in sheet)"
        route = f"⚠ VERIFY: /<section>/{slug}"
        schema_rec = ("Use FAQPage schema for this section, combined with MedicalWebPage and "
                      "BreadcrumbList schema for the parent page.")
    is_done = page["status"].lower() == "done"
    section_key = "sectionToMerge" if is_done else "section"
    out = {
        "pillar": page["title"],
        "contentType": page["content_type"] or "⚠ VERIFY",
        "runner": runner,
        "slug": f"⚠ VERIFY: {slug}",
        "route": route,
        section_key: section,
        "schemaRecommendation": schema_rec,
        "medicalDisclaimer": MEDICAL_DISCLAIMER,
    }
    return out


def main():
    ap = argparse.ArgumentParser(description="Batch FAQ-section generator for CancerFax content.")
    ap.add_argument("--xlsx", default=str(DEFAULT_XLSX))
    ap.add_argument("--sheet", default=DEFAULT_SHEET)
    ap.add_argument("--status", default="Pending", help="Filter rows by Status (e.g. Pending, Done, all)")
    ap.add_argument("--limit", type=int, default=2, help="Max rows this run (0 = no limit)")
    ap.add_argument("--model", default=DEFAULT_MODEL)
    ap.add_argument("--out", default=str(DEFAULT_OUT))
    ap.add_argument("--dry-run", action="store_true", help="List selected rows; no API calls, no writes")
    args = ap.parse_args()

    rows = read_rows(args.xlsx, args.sheet)
    if args.status.lower() != "all":
        rows = [r for r in rows if r["status"].lower() == args.status.lower()]
    if args.limit and args.limit > 0:
        rows = rows[: args.limit]

    if not rows:
        print(f"No rows matched status='{args.status}' in sheet '{args.sheet}'.")
        return 1

    print(f"Selected {len(rows)} row(s) (status='{args.status}', sheet='{args.sheet}'):")
    for r in rows:
        print(f"  row {r['row']}: {r['title']}  [pillar={r['pillar_name'] or '-'}, "
              f"type={r['content_type'] or '-'}, status={r['status'] or '-'}]")

    if args.dry_run:
        return 0

    import anthropic  # imported here so --dry-run works without the SDK/auth
    client = anthropic.Anthropic()
    master_prompt = PROMPT_PATH.read_text()
    out_dir = Path(args.out)
    out_dir.mkdir(parents=True, exist_ok=True)

    failures = 0
    for r in rows:
        slug = slugify(r["title"])
        targets = page_targets(r["content_type"])
        print(f"\n=== {r['title']} (row {r['row']}) ===")
        print(f"  page type: {r['content_type'] or 'blank->pillar default'} "
              f"-> target ~{targets['count']} FAQs / {targets['groups']} groups")
        try:
            print("  researching...")
            notes = research(client, r, args.model)
            print(f"  research notes: {len(notes)} chars")
            print("  generating...")
            section = generate_section(client, r, notes, master_prompt, targets, args.model)
            problems, warnings, count = validate(section, targets)
            if problems:
                print("  RETRY (validation problems): " + "; ".join(problems))
                # one corrective retry
                section = generate_section(
                    client, r,
                    notes + "\n\nThe previous attempt had these problems; fix them: " + "; ".join(problems),
                    master_prompt, targets, args.model,
                )
                problems, warnings, count = validate(section, targets)
            if problems:
                print("  FAILED validation: " + "; ".join(problems))
                failures += 1
                continue
            for w in warnings:
                print(f"  warning: {w}")
            out = build_output(r, section)
            path = out_dir / f"{slug}-faq-section.json"
            path.write_text(json.dumps(out, indent=2, ensure_ascii=False) + "\n")
            print(f"  wrote {path}  ({count} FAQs, {len(section['groups'])} groups)")
        except Exception as e:  # noqa: BLE001 - report and continue the batch
            print(f"  ERROR: {type(e).__name__}: {e}")
            failures += 1

    print(f"\nDone. {len(rows) - failures}/{len(rows)} succeeded.")
    return 1 if failures else 0


if __name__ == "__main__":
    sys.exit(main())
